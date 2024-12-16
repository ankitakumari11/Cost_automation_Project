from django.shortcuts import render, redirect ,get_object_or_404
from django.http import JsonResponse
from .models import CostCalculation, ScopingForm , CostSummary ,RateCard, Travel, PrivatePublicCloudProjectCost, Tool, On_Call, CostCalculation ,YearlyCostSummary ,ProjectResourceUtilisation
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from django.http import HttpResponse
from datetime import datetime
from decimal import Decimal
from django.contrib.auth import authenticate, login ,logout
from django.contrib import messages
from django.views.decorators.cache import cache_control
from django.db import connection
from django.apps import apps
from django.contrib.admin.sites import site
import csv

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def login_view(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Authenticate the user
        user = authenticate(request, username=username, password=password)
        if user is not None:
            # Log in the user
            login(request, user)
            return redirect('home')  # Replace 'home' with the name of your home page URL
        else:
            messages.error(request, "Invalid username or password")
    
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect ('login')

def export_all_data(request):
    # Fetch the latest ScopingForm instance
    scoping_form = ScopingForm.objects.last()  # Change this logic as per your requirements
    customer_name = scoping_form.customer_name if scoping_form else "default"
    project_name = scoping_form.project_name if scoping_form else "default"

    # If customer name is missing, return an error response
    if not project_name:
        return JsonResponse({"status": "error", "message": "Export failed: project_name is missing."})

    # Create a new workbook
    workbook = openpyxl.Workbook()
    del workbook['Sheet']  # Remove the default sheet

    # Define a border style
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # Add ScopingForm data in a new sheet (transposed)
    if scoping_form:
        scoping_form_sheet = workbook.create_sheet(title="ScopingForm")
        
        # Write field names in the first column and corresponding values in the second column (transpose)
        scoping_form_fields = [field.name for field in ScopingForm._meta.fields if field.name != "id"]  # Exclude 'id'
        
        for row_num, field_name in enumerate(scoping_form_fields, start=1):
            scoping_form_sheet[f"A{row_num}"] = field_name  # Field name in the first column
            value = getattr(scoping_form, field_name)
            scoping_form_sheet[f"B{row_num}"] = value  # Corresponding value in the second column
            # Apply border to each cell
            scoping_form_sheet[f"A{row_num}"].border = border
            scoping_form_sheet[f"B{row_num}"].border = border

    # Add data for each table
    model_sheets = {
        "CostCalculation": CostCalculation,
        "PrivatePublicCloudProjectCost": PrivatePublicCloudProjectCost,
        "Tools": Tool,
        "RateCard": RateCard,
        "Travel": Travel,      
        "On_Call": On_Call,
        
    }

    for sheet_name, model in model_sheets.items():
        sheet = workbook.create_sheet(title=sheet_name)

        # Get all objects and their fields
        objects = model.objects.all()
        if objects.exists():
            fields = [field.name for field in model._meta.fields if field.name != "id"]  # Exclude 'id'

            # Write headers and apply borders
            for col_num, field_name in enumerate(fields, start=1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}1"] = field_name
                sheet[f"{col_letter}1"].border = border  # Apply border to header

            # Write rows and apply borders
            for row_num, obj in enumerate(objects, start=2):
                for col_num, field_name in enumerate(fields, start=1):
                    col_letter = get_column_letter(col_num)
                    value = getattr(obj, field_name)
                    cell = sheet[f"{col_letter}{row_num}"]
                    cell.value = value
                    cell.border = border  # Apply border to each cell

    

    # Use customer_name in the filename, ensuring it's safe for a file name
    sanitized_project_name = "".join(c if c.isalnum() or c in (' ', '_') else "_" for c in project_name).strip()
    filename = f"{sanitized_project_name}_data.xlsx"

    # Save workbook to response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename={filename}"
    workbook.save(response)
    return response


def implement(request):
    # Just render the home page with the buttons
    return render(request, 'Implementation.html')



def export_report_table(request):
    table_id = request.GET.get('table_id')
    if not table_id:
        return HttpResponse("No table selected for export.", status=400)

    table_mapping = {
        "costSummary": "Managed_services_costsummary",
        "projectResourceUtilisation": "Managed_services_projectresourceutilisation",
        # "scopingForms": "Managed_services_scopingform",
        "yearlyCostSummary": "Managed_services_yearlycostsummary",
    }

    table_name = table_mapping.get(table_id)
    if not table_name:
        return HttpResponse("Invalid table selected.", status=400)

    # Query data from the database
    with connection.cursor() as cursor:
        cursor.execute(f"SELECT * FROM {table_name}")
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    # Create the response
    response = HttpResponse(content_type="text/csv")
    response['Content-Disposition'] = f'attachment; filename="{table_id}.csv"'

    # Write data to CSV
    writer = csv.writer(response)
    writer.writerow(columns)  # Write header
    writer.writerows(rows)    # Write data rows

    return response

def get_column_order(model_name):
    """Fetch column order based on `list_display` in admin.py."""
    model = apps.get_model("Managed_services", model_name)  # Replace with your app name
    admin_class = site._registry.get(model)  # Retrieve the admin class for the model
    if admin_class and hasattr(admin_class, "list_display"):
        return admin_class.list_display  # This returns the column order
    return None

def reports(request):
    def fetch_data(query, model_name=None):
        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

        # Get column order dynamically from `admin.py`
        if model_name:
            column_order = get_column_order(model_name)
            if column_order:
                column_mapping = {col: idx for idx, col in enumerate(columns)}
                reordered_columns = [col for col in column_order if col in columns]
                reordered_rows = [
                    [row[column_mapping[col]] for col in reordered_columns]
                    for row in rows
                ]
                return {"columns": reordered_columns, "rows": reordered_rows}

        return {"columns": columns, "rows": rows}

    cost_summary = fetch_data("SELECT * FROM Managed_services_costsummary", "CostSummary")
    project_resource_utilisation = fetch_data(
    "SELECT * FROM Managed_services_projectresourceutilisation", "ProjectResourceUtilisation"
)
    # scoping_forms = fetch_data("SELECT * FROM Managed_services_scopingform", "ScopingForm")
    yearly_cost_summary = fetch_data(
    "SELECT * FROM Managed_services_yearlycostsummary", "YearlyCostSummary"
)

    return render(request, "report.html", {
        "cost_summary": cost_summary,
        "project_resource_utilisation": project_resource_utilisation,
        # "scoping_forms": scoping_forms,
        "yearly_cost_summary": yearly_cost_summary,
    })


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def home(request):
     # Ensure the browser doesn't cache this page
    if not request.user.is_authenticated:
        return redirect('login') 
    # Just render the home page with the buttons
    return render(request, 'home.html')

# @cache_control(no_cache=True, must_revalidate=True, no_store=True)
def modify_form(request):
    if request.method == 'GET':
        projects = ScopingForm.objects.all()
        return render(request, 'modify_form.html', {'projects': projects})
    
    elif request.method == 'POST':
        Project_Name = request.POST.get('ProjectName')
        project = get_object_or_404(ScopingForm, project_name=Project_Name)
        
        # Update project with new values
        project.SA_name=request.user.username
        project.customer_name = request.POST.get('CustomerName')
        project.product_name = request.POST.get('ProductName')
        project.description = request.POST.get('Description', "No description provided")
        project.location = request.POST.get('Location', "Not specified")

        # Convert received_date and submission_date to datetime
        received_date = request.POST.get('ReceivedDate')
        if received_date:
            project.received_date = datetime.strptime(received_date, '%Y-%m-%d')

        submission_date = request.POST.get('SubmissionDate')
        if submission_date:
            project.submission_date = datetime.strptime(submission_date, '%Y-%m-%d')

        project.project_status = request.POST.get('ProjectStatus')

        # Convert margin and penalty_risk to Decimal
        project.margin = Decimal(request.POST.get('Margin', 0))
        project.penalty_risk = Decimal(request.POST.get('PenaltyRisk', 0))

        # Convert numerical fields to integers
        project.no_of_windows_vms = int(request.POST.get('WindowsVM', 0))
        project.no_of_linux_vms = int(request.POST.get('LinuxVM', 0))
        project.no_of_rdbms_dbs = int(request.POST.get('RDBMdB', 0))
        project.no_of_nosql_dbs = int(request.POST.get('NOSQLdB', 0))
        project.no_of_network_devices = int(request.POST.get('NetworkDevices', 0))
        project.Total_VMs_Devices = int(request.POST.get('Total_VMs_Devices', 0))
        project.no_of_ad_objects = int(request.POST.get('noOfAD', 0))
        project.contract_duration = int(request.POST.get('ContractDuration', 0))
        project.no_of_private_cloud_hosts = int(request.POST.get('PriavteCloudHosts', 0))
        project.no_of_servers_for_dr = int(request.POST.get('ServersforDR', 0))
        project.dr_drills = int(request.POST.get('DRdrills', 0))

        # Convert total_storage_capacity and yoy_increment to Decimal
        project.total_storage_capacity = Decimal(request.POST.get('StorageCapacity', 0))
        project.yoy_increment = Decimal(request.POST.get('YoYIncrement', 0))

        project.on_call = request.POST.get('OnCall')
        project.support_window = request.POST.get('SupportWindow')
        project.dc_type = request.POST.get('TypeOfDC')
        project.hyper_converged_platform_used = request.POST.get('ConvergedPlatform')
        project.dr_in_scope = request.POST.get('DRInScope')
        project.complexity = request.POST.get('Complexity')
        project.monitoring = request.POST.get('Monitoring')
        project.patching = request.POST.get('Patching')
        project.itsm_services = request.POST.get('ITSMservices')
        project.ipc_management_in_scope = request.POST.get('IPCManagement')
        project.travel = request.POST.get('Travel')
        project.management_governance_support = request.POST.get('GovernanceSupport')

        # Convert tcv to Decimal
        project.tcv = Decimal(request.POST.get('TCV', 0))

        project.save()

        # Trigger calculation of final cost
        final_cost = CostCalculation.calculate_values()

        product_name_sf=request.POST.get('ProductName')
        # Filter the corresponding CostSummary entry based on project name
        cost_summary = CostSummary.objects.filter(project_name=Project_Name).first()

        
        return render(request, 'modify_form.html', {'final_cost': final_cost, 'cost_summary': cost_summary ,'projects': ScopingForm.objects.all(),'product_name':product_name_sf, 'message': 'Project updated successfully!'})

def fetch_project_data(request):
    # Get project_name from the query parameters
    project_name = request.GET.get('project_name')  # Ensure consistency in parameter name

    # Retrieve the project using project_name
    project = get_object_or_404(ScopingForm, project_name=project_name)
    
    # Prepare the data to send as JSON
    data = {
        # 'SA_name':project.SA_name,
        'CustomerName': project.customer_name,
        'product_name': project.product_name,
        'Description': project.description,
        'Location': project.location,
        'ReceivedDate': project.received_date,
        'SubmissionDate': project.submission_date,
        'ProjectStatus': project.project_status,
        'Margin': project.margin,
        'PenaltyRisk': project.penalty_risk,
        'WindowsVM': project.no_of_windows_vms,
        'LinuxVM': project.no_of_linux_vms,
        'RDBMdB': project.no_of_rdbms_dbs,
        'NOSQLdB': project.no_of_nosql_dbs,
        'NetworkDevices': project.no_of_network_devices,
        'Total_VMs_Devices': project.Total_VMs_Devices,
        'StorageCapacity': project.total_storage_capacity,
        'OnCall': project.on_call,
        'SupportWindow': project.support_window,
        'YoYIncrement': project.yoy_increment,
        'patching': project.patching,
        'noOfAD': project.no_of_ad_objects,
        'ContractDuration': project.contract_duration,
        'TypeOfDC': project.dc_type,
        'PriavteCloudHosts': project.no_of_private_cloud_hosts,
        'ConvergedPlatform': project.hyper_converged_platform_used,
        'dr_in_scope': project.dr_in_scope,
        'no_of_servers_for_dr':project.no_of_servers_for_dr,
        'dr_drills':project.dr_drills,
        'complexity':project.complexity,
        'monitoring':project.monitoring,
        'patching':project.patching,
        'itsm_services':project.itsm_services,
        'ipc_management_in_scope':project.ipc_management_in_scope,
        'travel':project.travel,
        'management_governance_support':project.management_governance_support,
        'tcv':project.tcv,

    }
    # print(f"Received project_name: {project_name}")
    # print(f"Data sent: {data}")
    # Return the data as a JSON response
    return JsonResponse(data)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def scoping_form(request):
    # Check if the form has been submitted (POST request)
    if request.method == 'POST':
        # Retrieve form data from the POST request
        customer_name = request.POST.get('CustomerName')
        project_name = request.POST.get('ProjectName')

        # Check if the ScopingForm already exists to avoid duplicates
        existing_form = ScopingForm.objects.filter(project_name=project_name).exists()

        if not existing_form:  # Only create a new instance if it doesn't exist
            try:
                # Create the new ScopingForm instance
                scoping_form_instance = ScopingForm(
                    # Save the logged-in user's username
                    SA_name=request.user.username,
                    customer_name=customer_name,
                    project_name=project_name,
                    description=request.POST.get('Description'),
                    product_name=request.POST.get('ProductName'),
                    location=request.POST.get('Location'),
                    received_date=request.POST.get('ReceivedDate'),
                    submission_date=request.POST.get('SubmissionDate'),
                    project_status=request.POST.get('ProjectStatus'),
                    margin=request.POST.get('Margin'),
                    penalty_risk=request.POST.get('PenaltyRisk'),
                    no_of_windows_vms=request.POST.get('WindowsVM'),
                    no_of_linux_vms=request.POST.get('LinuxVM'),
                    no_of_rdbms_dbs=request.POST.get('RDBMdB'),
                    no_of_nosql_dbs=request.POST.get('NOSQLdB'),
                    no_of_network_devices=request.POST.get('NetworkDevices'),
                    Total_VMs_Devices=request.POST.get('Total_VMs_Devices'),
                    total_storage_capacity=request.POST.get('StorageCapacity'),
                    on_call=request.POST.get('OnCall'),
                    support_window=request.POST.get('SupportWindow'),
                    yoy_increment=request.POST.get('YoYIncrement'),
                    no_of_ad_objects=request.POST.get('noOfAD'),
                    contract_duration=request.POST.get('ContractDuration'),
                    dc_type=request.POST.get('TypeOfDC'),
                    no_of_private_cloud_hosts=request.POST.get('PriavteCloudHosts'),
                    hyper_converged_platform_used=request.POST.get('ConvergedPlatform'),
                    dr_in_scope=request.POST.get('DRInScope'),
                    no_of_servers_for_dr=request.POST.get('ServersforDR'),
                    dr_drills=request.POST.get('DRdrills'),
                    complexity=request.POST.get('Complexity'),
                    monitoring=request.POST.get('Monitoring'),
                    patching=request.POST.get('Patching'),
                    itsm_services=request.POST.get('ITSMservices'),
                    ipc_management_in_scope=request.POST.get('IPCManagement'),
                    travel=request.POST.get('Travel'),
                    management_governance_support=request.POST.get('GovernanceSupport'),
                    tcv=request.POST.get('TCV'),
                    
                )
                scoping_form_instance.save()

                # Trigger calculation of final cost
                final_cost = CostCalculation.calculate_values()

                # Get the last CostSummary entry
                cost_summary = CostSummary.objects.last()

                # Redirect to the same page or another page (e.g., success page)
                return render(request, 'form.html', {'final_cost': final_cost, 'cost_summary': cost_summary})

            except Exception as e:
                # Handle unexpected errors during form save
                return render(request, 'form.html', {'error': f"An error occurred: {str(e)}"})

        else:
            
            # If the form already exists, notify the user
            error_message = f"A project with the name '{project_name}' already exists in the database. Please use a different name."
            print(error_message)
            return render(request, 'form.html', {
                'error_message': error_message,
            })

    else:

        # Reset final_cost and cost_summary to None on initial load or page reload
        final_cost = None
        cost_summary = None

        return render(request, 'form.html', {
            'final_cost': final_cost,
            'cost_summary': cost_summary,
            
        })







